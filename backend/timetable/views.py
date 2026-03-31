from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import FileResponse
from .models import Faculty, Subject, Section, Timetable
from .serializers import (
    FacultySerializer, SubjectSerializer, SectionSerializer, 
    TimetableSerializer, TimetableGenerationSerializer
)
from .utils import TimetableGenerator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

class FacultyViewSet(viewsets.ModelViewSet):
    """Faculty management"""
    queryset = Faculty.objects.all()
    serializer_class = FacultySerializer
    permission_classes = [IsAuthenticated]
    search_fields = ['name', 'email', 'department']
    ordering_fields = ['name', 'created_at']

class SubjectViewSet(viewsets.ModelViewSet):
    """Subject management"""
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ['name', 'code']
    ordering_fields = ['code', 'name', 'created_at']

class SectionViewSet(viewsets.ModelViewSet):
    """Section management"""
    queryset = Section.objects.all()
    serializer_class = SectionSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ['name']
    ordering_fields = ['name', 'semester', 'created_at']

class TimetableViewSet(viewsets.ModelViewSet):
    """Timetable management"""
    queryset = Timetable.objects.all()
    serializer_class = TimetableSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ['section__name', 'subject__code', 'faculty__name']
    ordering_fields = ['day', 'time_slot', 'created_at']
    
    def get_queryset(self):
        """Filter by section if provided"""
        queryset = Timetable.objects.all()
        section_id = self.request.query_params.get('section_id')
        
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def generate(self, request):
        """
        Generate timetable for a section
        
        POST /api/timetable/timetable/generate/
        {
            "section_id": 1
        }
        """
        serializer = TimetableGenerationSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        section_id = serializer.validated_data['section_id']
        section = Section.objects.get(id=section_id)
        
        # Delete existing timetable for this section
        Timetable.objects.filter(section=section).delete()
        
        # Generate new timetable
        try:
            generator = TimetableGenerator(section)
            result = generator.generate()
            
            return Response({
                'status': 'success',
                'message': f"Generated {len(result['schedule'])} timetable slots",
                'total_scheduled': result['total_scheduled'],
                'conflicts': result['conflicts'],
                'timetable': TimetableSerializer(result['schedule'], many=True).data
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export timetable to Excel
        
        GET /api/timetable/timetable/export/?section_id=1
        """
        section_id = request.query_params.get('section_id')
        
        if not section_id:
            return Response(
                {'error': 'section_id parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            section = Section.objects.get(id=section_id)
            timetables = Timetable.objects.filter(section=section).order_by('day', 'time_slot')
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Timetable - {section.name}"
            
            # Add title
            ws['A1'] = f"Timetable for {section.name}"
            ws['A1'].font = Font(bold=True, size=14)
            
            # Add headers
            headers = ['Day', 'Time Slot', 'Subject', 'Faculty', 'Room', 'Type']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data
            row = 4
            for timetable in timetables:
                ws.cell(row=row, column=1).value = timetable.get_day_display()
                ws.cell(row=row, column=2).value = timetable.time_slot
                ws.cell(row=row, column=3).value = timetable.subject.name
                ws.cell(row=row, column=4).value = timetable.faculty.name if timetable.faculty else 'TBD'
                ws.cell(row=row, column=5).value = timetable.room
                ws.cell(row=row, column=6).value = timetable.subject.get_subject_type_display()
                row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            
            # Save to bytes
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return FileResponse(
                excel_file,
                as_attachment=True,
                filename=f"timetable_{section.name}_{request.build_absolute_uri().split('/')[-1]}.xlsx"
            )
        
        except Section.DoesNotExist:
            return Response(
                {'error': 'Section not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
